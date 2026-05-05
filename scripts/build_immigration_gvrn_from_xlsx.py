# -*- coding: utf-8 -*-
"""Build SAMPLE_BY_REGION from 관할_출입국외국인관서_조회.xlsx and patch platform-faq.html."""
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
FAQ_PATH = ROOT / "platform-faq.html"
DATA_JSON = ROOT / "data" / "immigration-gvrn-by-region.json"

REGION_ORDER = [
    "서울",
    "경기",
    "인천",
    "강원",
    "충남",
    "대전",
    "충북",
    "부산",
    "울산",
    "대구",
    "경북",
    "경남",
    "전남",
    "광주",
    "전북",
    "제주",
]

SEARCH_BOOST_BY_NAME = {
    "부산출입국·외국인청": (
        "부산시 부산광역시 중구 서구 동구 영도구 부산진구 동래구 남구 북구 해운대구 사하구 금정구 "
        "강서구 연제구 수영구 사상구 기장군 행정구 일반 체류 민원"
    ),
}


def clean_jurisdiction(text: str) -> str:
    if not text:
        return text
    # 한글·숫자·닫는 괄호 바로 뒤에 붙은 ■ 앞에 공백 (예: 인천광역시■ 경기도)
    text = re.sub(r"([가-힣0-9\)])■", r"\1 ■", text)
    return text.strip()


def load_from_xlsx(path: Path) -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    by_region: dict[str, list[dict]] = defaultdict(list)
    for r in rows[1:]:
        if not r or all(x is None or str(x).strip() == "" for x in r):
            continue
        pad = list(r) + [None] * 4
        region = str(pad[0]).strip() if pad[0] is not None else ""
        name = str(pad[1]).strip() if pad[1] is not None else ""
        addr = str(pad[2]).strip() if pad[2] is not None else ""
        jur = clean_jurisdiction(str(pad[3]).strip() if pad[3] is not None else "")
        if not region or not name:
            continue
        row_obj: dict = {"name": name, "address": addr, "jurisdiction": jur}
        boost = SEARCH_BOOST_BY_NAME.get(name)
        if boost:
            row_obj["searchBoost"] = boost
        by_region[region].append(row_obj)

    ordered: dict[str, list[dict]] = {}
    for key in REGION_ORDER:
        if key in by_region:
            ordered[key] = sorted(by_region[key], key=lambda x: x["name"])
    for key in sorted(by_region.keys()):
        if key not in ordered:
            ordered[key] = sorted(by_region[key], key=lambda x: x["name"])

    return ordered


def strip_for_dump(row: dict) -> dict:
    return {k: v for k, v in row.items() if v is not None and v != ""}


def format_sample_assignment(indent: str, data: dict[str, list[dict]]) -> str:
    dumped = {
        region: [strip_for_dump(o) for o in offices] for region, offices in data.items()
    }
    body = json.dumps(dumped, ensure_ascii=False, indent=2)
    lines = body.split("\n")
    out = [indent + "var SAMPLE_BY_REGION = " + lines[0]]
    out.extend(indent + line for line in lines[1:])
    return "\n".join(out) + ";"


def replace_sample_block(html: str, data: dict[str, list[dict]]) -> str:
    needle = "var SAMPLE_BY_REGION = "
    start = html.find(needle)
    if start == -1:
        raise SystemExit("platform-faq.html: SAMPLE_BY_REGION not found")

    line_start = html.rfind("\n", 0, start) + 1
    indent = "      "

    brace_start = html.find("{", start + len(needle))
    if brace_start == -1:
        raise SystemExit("platform-faq.html: opening brace not found")

    depth = 0
    in_str = False
    esc = False
    quote = ""
    i = brace_start
    while i < len(html):
        ch = html[i]
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == quote:
                in_str = False
            i += 1
            continue
        if ch in "\"'":
            in_str = True
            quote = ch
            i += 1
            continue
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                end_obj = i + 1
                if end_obj < len(html) and html[end_obj] == ";":
                    end_obj += 1
                new_assignment = format_sample_assignment(indent, data)
                return html[:line_start] + new_assignment + html[end_obj:]
        i += 1

    raise SystemExit("platform-faq.html: unbalanced braces in SAMPLE_BY_REGION")


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads" / "관할_출입국외국인관서_조회.xlsx"
    if not xlsx.is_file():
        raise SystemExit(f"Excel not found: {xlsx}")

    data = load_from_xlsx(xlsx)
    DATA_JSON.parent.mkdir(parents=True, exist_ok=True)
    DATA_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    html = FAQ_PATH.read_text(encoding="utf-8")
    html_new = replace_sample_block(html, data)

    comment_needle = "「관할 출입국·외국인관서 조회」"
    if comment_needle in html_new and "관할_출입국외국인관서_조회.xlsx" not in html_new:
        html_new = html_new.replace(
            "/** 하이코리아 「관할 출입국·외국인관서 조회」 목록 발췌 — 전체·최신 정보는 공식 페이지 우선 */",
            "/** 하이코리아 「관할 출입국·외국인관서 조회」 — 관할_출입국외국인관서_조회.xlsx 반영, "
            "전체·최신 정보는 공식 페이지 우선 */",
            1,
        )

    FAQ_PATH.write_text(html_new, encoding="utf-8")

    total = sum(len(v) for v in data.values())
    print(json.dumps({"xlsx": str(xlsx), "regions": len(data), "offices": total, "json": str(DATA_JSON)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
