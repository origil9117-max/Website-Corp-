# -*- coding: utf-8 -*-
"""Build embassy-groups.json array from 재외공관정보.xlsx."""
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

SLUG_RE = re.compile(r"mofa\.go\.kr/([^/]+)/?", re.I)


def slug_to_base(slug: str) -> str:
    s = slug.strip().lower()
    if s.endswith("-ko"):
        s = s[:-3]
    return s.split("-")[0]


def place_from_mission_name(name: str) -> str:
    name = (name or "").strip()
    if "대한민국" in name and name.startswith("주"):
        i = name.index("대한민국")
        return name[1:i].strip()
    if name.startswith("한-"):
        return "아세안·국제기구"
    return name


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads" / "재외공관정보.xlsx"
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    missions_flat = []
    for r in rows[1:]:
        if not r or all(x is None or str(x).strip() == "" for x in r):
            continue
        pad = list(r) + [None] * (7 - len(r))

        def cell(i):
            v = pad[i]
            if v is None:
                return None
            if isinstance(v, float) and v == int(v):
                v = int(v)
            return str(v).strip() if isinstance(v, str) else v

        name = cell(1)
        url = cell(6)
        if not name or not url:
            continue

        m = SLUG_RE.search(str(url))
        if not m:
            print("WARN no slug:", name, url, file=sys.stderr)
            continue
        slug_full = m.group(1)
        base = slug_to_base(slug_full)

        phone = cell(2) or ""
        fax = cell(3) or ""
        zipc = cell(4)
        addr = cell(5) or ""

        phone_str = str(phone) if phone else ""
        if fax:
            phone_str += (" / 팩스: " + str(fax)) if phone_str else ("팩스: " + str(fax))
        if not phone_str.strip():
            phone_str = "—"

        address = str(addr) if addr else "—"
        if zipc and str(zipc) and str(zipc) not in address:
            address = f"(우편) {zipc} · {address}"

        missions_flat.append(
            {
                "name": str(name),
                "phone": phone_str,
                "address": address,
                "homepage": str(url).strip(),
                "_slug": slug_full,
                "_base": base,
            }
        )

    by_base: dict[str, list] = defaultdict(list)
    for row in missions_flat:
        by_base[row["_base"]].append(row)

    def label_for_base(base: str, ms: list) -> str:
        preferred = None
        for row in ms:
            if row["_slug"].lower() == f"{base}-ko":
                preferred = row
                break
        pick = preferred or sorted(ms, key=lambda x: x["name"])[0]
        tok = place_from_mission_name(pick["name"])
        aliases = {"호주연방": "호주"}
        return aliases.get(tok, tok)

    groups = []
    for base in sorted(by_base.keys()):
        ms = by_base[base]
        label = label_for_base(base, ms)
        missions = []
        for row in sorted(ms, key=lambda x: x["name"]):
            missions.append(
                {
                    "name": row["name"],
                    "phone": row["phone"],
                    "address": row["address"],
                    "homepage": row["homepage"],
                }
            )
        groups.append({"base": base, "label": label, "missions": missions})

    groups.sort(key=lambda g: g["label"].replace(" ", ""))

    root = Path(__file__).resolve().parent.parent / "data"
    root.mkdir(parents=True, exist_ok=True)
    meta = {"missions": len(missions_flat), "countries": len(groups), "bases": [g["base"] for g in groups]}
    (root / "embassy-built-meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    out_path = root / "embassy-groups.json"
    out_path.write_text(json.dumps(groups, ensure_ascii=False), encoding="utf-8")

    js_path = root / "embassy-data.js"
    js_path.write_text(
        "window.EMBASSY_GROUPS = "
        + json.dumps(groups, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )

    print(json.dumps({"json": str(out_path), "js": str(js_path), **meta}, ensure_ascii=False))


if __name__ == "__main__":
    main()
